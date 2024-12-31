from customtkinter import *
from datetime import datetime
import popup
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image


def number_to_excel_column(n):
    result = ""
    while n >= 0:
        result = chr(n % 26 + ord('A')) + result
        n = n // 26 - 1
    return result

class create_member:
    def __init__(self, data):
        self.picture = data[0]

        self.prenom1 = data[1]
        self.prenom2 = data[2]
        self.prenom3 = data[3]

        self.nom = data[4]
        self.nom_usage = data[5]

        self.metier = data[6]
        self.sexe = data[7]

        self.date_n1 = data[8]
        self.date_n2 = data[9]
        self.date_n3 = data[10]
        self.commune_n = data[11]
        self.departement_n = data[12]

        self.date_m1 = data[13]
        self.date_m2 = data[14]
        self.date_m3 = data[15]
        self.commune_m = data[16]
        self.departement_m = data[17]

        self.date_d1 = data[18]
        self.date_d2 = data[19]
        self.date_d3 = data[20]
        self.commune_d = data[21]
        self.departement_d = data[22]

        self.pere = data[23]
        self.mere = data[24]
        self.conjoint = data[25]

        self.completion = int(data[26])
        self.date = data[27]   #(datetime.now()).strftime("Le %d/%m/%Y à %H:%M")
        
        self.column = data[-1] #une sorte d'index

        self.notes_l = []
        self.photo_l = []
        for i in range(28,len(data)-1):
            if data[i] in self.column:
                self.photo_l.append(data[i])
            else:
                self.notes_l.append((data[i].split("\n"))[0])
                self.notes_l.append(data[i])

def open_member(self,event):
    if not self.root.opened and not self.root.swt_edit.winfo_exists():
        self.selected_mb = self.root.selected_mb
        self.files_add = []
        self.files_supp = []
        self.sld_file = (None,None)
        self.id_pere = self.selected_mb.pere
        self.id_mere = self.selected_mb.mere
        self.id_conjoint = self.selected_mb.conjoint
        if not self.selected_mb.picture is None:
            self.photo_location = f"{self.root.directory_path}\\temp\\{self.selected_mb.picture}.png"
        else:
            self.photo_location = self.root.directory_path + "\\assets\\default.jpg"

        asign = {"prenom1":self.ntr_prenom1, "prenom2":self.ntr_prenom2, "prenom3":self.ntr_prenom3, "nom":self.cbb_nom, "nom_usage":self.cbb_nom_usage, "metier":self.ntr_metier, "sexe":self.sexe_var, "date_n1":self.ntr_dn1, "date_n2":self.ntr_dn2, "date_n3":self.ntr_dn3, "commune_n":self.ntr_cn, "departement_n":self.btn_dn, "date_m1":self.ntr_dn1, "date_m2":self.ntr_dn2, "date_m3":self.ntr_dn3, "commune_m":self.ntr_cn, "departement_m":self.btn_dn, "date_d1":self.ntr_dn1, "date_d2":self.ntr_dn2, "date_d3":self.ntr_dn3, "commune_d":self.ntr_cn, "departement_d":self.btn_dn, "pere":self.btn_pere, "mere":self.btn_mere, "conjoint":self.btn_conjoint, "completion":self.pbar_comple, "date":None, "picture":None, "column": None,"notes_l":None,"photo_l":None}
        data = vars(self.selected_mb)

        for key,val in data.items():
            widget = asign[key]
            if isinstance(widget, CTkEntry):
                widget.delete(0, "end")
                if val not in ("Inconnu","??","????"):
                    widget.insert(0, val)  
                else:
                    widget.focus()
                    self.update_idletasks()

            elif isinstance(widget, CTkButton):
                if key in ("pere","mere","conjoint"):
                    self.id_to_text(widget,val)

                else:
                    widget.configure(text=val)
                    
                if val == "disabled":
                    for index,btn in enumerate(("conjoint", "departement_m", "departement_d")):
                        if key == btn:
                            self.checkbox_action(index)

            elif isinstance(widget, (CTkComboBox, CTkProgressBar)):
                widget.set(val if key != "completion" else val/100)
                
            elif key == 'picture':
                self.select_photo(self.photo_location)

            elif key == "sexe":
                self.sexe_var = val
                if val == "homme":
                    self.rbtn_homme.select()
                else:
                    self.rbtn_femme.select()

            elif key == "photo_l":
                for photos in val:
                    photo = CTkFrame(self.sfrm_file, text=photos)
                    photo.pack()

            elif key == "notes_l":
                for i in range(int(len(val)/2)):
                    note = CTkFrame(self.sfrm_file, text=val[i*2])
                    note.pack()

        self.modifications = 0
        self.root.opened = True
        show_files(self)

def save(self):
    dict_ = get_dict(self)
    matcher = {"nom":"5","nom_usage":"6","date_n1":"9","date_n2":"10","date_n3":"11","prenom1": "2","prenom2": "3","prenom3": "4","commune_n": "12","metier": "7","departement_n": "13","sexe": "8","pere": "24","mere": "25","conjoint":"26","date_m1":"14","date_m2":"15","date_m3":"16","commune_m":"17","departement_m":"18","date_d1":"19","date_d2":"20","date_d3":"21","commune_d":"22","departement_d":"23"}
    column = self.selected_mb.column

    wb = load_workbook(self.root.directory_path + "\\" + "data.xlsx")
    sheet = wb.active

    cell = column + "1"
    if not "default.jpg" in self.photo_location and not "temp" in self.photo_location:
    
        img = Image(self.photo_location)
        img.anchor = cell
        sheet.add_image(img)

    for key,val in dict_.items():
        if key != "picture" and key !="completion":
            row = matcher[key]
            sheet[column + row] = val

    date = (datetime.now()).strftime("Le %d/%m/%Y à %H:%M")
    sheet[column + "28"] = date

    self.check_completion()
    completion = (self.lbl_percent.cget("text")).replace("%","")
    sheet[column + "27"] = completion

    wb.save(self.root.directory_path + "\\" + "data.xlsx")

    self.modifications = 0
    self.change_dot() 

def get_dict(self):
    photo = os.path.basename(self.photo_location)
    widget_list = {"nom":self.cbb_nom.get(),"nom_usage": self.cbb_nom_usage.get(),"date_n1": self.ntr_dn1.get(),"date_n2": self.ntr_dn2.get(),"date_n3": self.ntr_dn3.get(),"prenom1": self.ntr_prenom1.get(),"prenom2": self.ntr_prenom2.get(),"prenom3": self.ntr_prenom3.get(),"commune_n": self.ntr_cn.get(),"metier": self.ntr_metier.get(),"departement_n": self.btn_dn.cget("text"),"sexe": self.sexe_var,"pere": self.id_pere,"mere": self.id_mere,"conjoint":"disabled","date_m1":"disabled","date_m2":"disabled","date_m3":"disabled","commune_m":"disabled","departement_m":"disabled","date_d1":"disabled","date_d2":"disabled","date_d3":"disabled","commune_d":"disabled","departement_d":"disabled","completion": self.pbar_comple.get()*100,"picture": photo if photo != "default.jpg" else None}

    if self.ckb_conjoint.get():
        widget_list["conjoint"] = self.id_conjoint

    if self.ckb_mariage.get():
        widget_list["date_m1"] = self.ntr_dm1.get()
        widget_list["date_m2"] = self.ntr_dm2.get()
        widget_list["date_m3"] = self.ntr_dm3.get()
        widget_list["commune_m"] = self.ntr_cm.get()
        widget_list["departement_m"] = self.btn_dm.cget("text")

    if self.ckb_deces.get():
        widget_list["date_d1"] = self.ntr_dd1.get()
        widget_list["date_d2"] = self.ntr_dd2.get()
        widget_list["date_d3"] = self.ntr_dd3.get()
        widget_list["commune_d"] = self.ntr_cd.get()
        widget_list["departement_d"] = self.btn_dd.cget("text")

    for key,value in widget_list.items():
        if isinstance(value, str) and (value.isspace() or value == ""):
            widget_list[key] = "Inconnu"

    return widget_list

def show_files(self):
    for widget in self.sfrm_file.grid_slaves():
        widget.destroy()

    index = 0
    for i in range(len(self.selected_mb.photo_l)):
        lbl_file = CTkLabel(self.sfrm_file, text="photo"+str(i), anchor="w",fg_color = "#333333")
        lbl_file.grid(row=index,column=0,padx=0,pady=0,sticky="ew")
        index += 1

    for j in range(int(len(self.selected_mb.notes_l)/2)):
        lbl_file = CTkLabel(self.sfrm_file, text=self.selected_mb.notes_l[j*2], anchor="w",fg_color = "#333333")
        lbl_file.grid(row=index,column=0,padx=0,pady=0,sticky="ew")
        index += 1

    for child in self.sfrm_file.grid_slaves():
        child.bind('<Enter>', lambda event: child.configure(fg_color = "grey"))
        child.bind('<Leave>', lambda event: child.configure(fg_color = "#333333"))
        child.bind("<Button-1>", self.choose_file)

    self.update_idletasks()
    self.sfrm_file.configure(width=max(300*self.root.scale, self.sfrm_file.winfo_reqwidth()))

def message(self,mess,type_,oui,func=None):
    title_ = ("Avertissement","Information")[type_]
    self.top_level = popup.C_popup(self,mess,title_,oui,func)

def importer(root):
    file_path = filedialog.askopenfilename(title="Ouvrir une feuille de calcul", filetypes=[("Feuille de calcul", "*.xlsx")])
    if file_path != "":
        wb = load_workbook(file_path)
        wbb = load_workbook(root.directory_path + "\\" + "data.xlsx")

        for sheet_name in wbb.sheetnames:
            del wbb[sheet_name]

        # Copier toutes les feuilles de calcul de wb dans wbb
        for sheet_name in wb.sheetnames:
            wbb.create_sheet(title=sheet_name)
            source_sheet = wb[sheet_name]
            dest_sheet = wbb[sheet_name]

            # Copier les données de la feuille de calcul
            for row in source_sheet.iter_rows(values_only=True):
                dest_sheet.append(row)

        wbb.save(root.directory_path + "\\" + "data.xlsx")

def clear(root):
    wb = load_workbook(root.directory_path + "\\" + "data.xlsx")

    for sheet_name in wb.sheetnames:
        del wb[sheet_name]

    wb.create_sheet(title="Sheet")
    wb.save(root.directory_path + "\\" + "data.xlsx")
